#!/usr/bin/env python3
"""
Add Industrial/Warehouse Stress Tab to Capital Stress Workbook
Per Derek's requirement: 19% of CRE exposure needs stress testing
"""

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Load existing workbook
workbook_path = 'evidence/capital_stress_2025Q2.xlsx'
wb = load_workbook(workbook_path)

# Create new sheet
ws = wb.create_sheet("Industrial_Warehouse_Stress")

# Header
ws['A1'] = 'INDUSTRIAL/WAREHOUSE STRESS SCENARIOS'
ws['A1'].font = Font(bold=True, size=14)
ws['A2'] = 'Exposure: $1.9B (19% of CRE) - Industrial $636M + Warehouse $1,295M'
ws['A2'].font = Font(italic=True)

# Exposure breakdown
row = 4
ws[f'A{row}'] = 'EXPOSURE BREAKDOWN'
ws[f'A{row}'].font = Font(bold=True, size=12)
row += 1

ws[f'A{row}'] = 'Property Type'
ws[f'B{row}'] = '$ Millions'
ws[f'C{row}'] = '% of CRE'
ws[f'D{row}'] = '% of Total Loans'
for col in [f'A{row}', f'B{row}', f'C{row}', f'D{row}']:
    ws[col].font = Font(bold=True)
    ws[col].fill = PatternFill(start_color='D4AF37', end_color='D4AF37', fill_type='solid')

row += 1
ws[f'A{row}'] = 'Warehouse'
ws[f'B{row}'] = 1295
ws[f'C{row}'] = '12.5%'
ws[f'D{row}'] = '6.5%'

row += 1
ws[f'A{row}'] = 'Industrial'
ws[f'B{row}'] = 636
ws[f'C{row}'] = '6.1%'
ws[f'D{row}'] = '3.2%'

row += 1
ws[f'A{row}'] = 'TOTAL INDUSTRIAL/WAREHOUSE'
ws[f'B{row}'] = 1931
ws[f'C{row}'] = '18.6%'
ws[f'D{row}'] = '9.8%'
ws[f'A{row}'].font = Font(bold=True)
ws[f'B{row}'].font = Font(bold=True)
ws[f'C{row}'].font = Font(bold=True)
ws[f'D{row}'].font = Font(bold=True)

# Base Case Scenario
row += 3
ws[f'A{row}'] = 'SCENARIO 1: BASE CASE (5% CUMULATIVE LOSS RATE)'
ws[f'A{row}'].font = Font(bold=True, size=12)
ws[f'A{row}'].fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
row += 1

headers = ['Metric', 'Formula/Value', 'Result', 'Units']
for col_idx, header in enumerate(headers, start=1):
    cell = ws.cell(row=row, column=col_idx)
    cell.value = header
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color='D4AF37', end_color='D4AF37', fill_type='solid')

row += 1
ws[f'A{row}'] = 'Exposure (M)'
ws[f'B{row}'] = '1,931'
ws[f'C{row}'] = 1931
ws[f'D{row}'] = 'millions'

row += 1
ws[f'A{row}'] = 'Loss Rate (%)'
ws[f'B{row}'] = '5%'
ws[f'C{row}'] = 0.05
ws[f'D{row}'] = 'decimal'

row += 1
ws[f'A{row}'] = 'Expected Loss (M)'
ws[f'B{row}'] = '=C' + str(row-2) + '*C' + str(row-1)
ws[f'C{row}'] = '=C' + str(row-2) + '*C' + str(row-1)
ws[f'D{row}'] = 'millions'

row += 1
ws[f'A{row}'] = 'Tax Rate'
ws[f'B{row}'] = '20%'
ws[f'C{row}'] = 0.20
ws[f'D{row}'] = 'decimal'

row += 1
ws[f'A{row}'] = 'After-Tax Impact (M)'
ws[f'B{row}'] = '=C' + str(row-2) + '*(1-C' + str(row-1) + ')'
ws[f'C{row}'] = '=C' + str(row-2) + '*(1-C' + str(row-1) + ')'
ws[f'D{row}'] = 'millions'

row += 1
ws[f'A{row}'] = 'RWA (M)'
ws[f'B{row}'] = '19,118.5'
ws[f'C{row}'] = 19118.5
ws[f'D{row}'] = 'millions'

row += 1
ws[f'A{row}'] = 'CET1 Burn (bps)'
ws[f'B{row}'] = '=(C' + str(row-2) + '/C' + str(row-1) + ')*10000'
ws[f'C{row}'] = '=(C' + str(row-2) + '/C' + str(row-1) + ')*10000'
ws[f'D{row}'] = 'bps'

row += 1
ws[f'A{row}'] = 'Starting CET1 Ratio (%)'
ws[f'B{row}'] = '13.35'
ws[f'C{row}'] = 13.35
ws[f'D{row}'] = 'percent'

row += 1
ws[f'A{row}'] = 'New CET1 Ratio (%)'
ws[f'B{row}'] = '=C' + str(row-1) + '-(C' + str(row-2) + '/100)'
ws[f'C{row}'] = '=C' + str(row-1) + '-(C' + str(row-2) + '/100)'
ws[f'D{row}'] = 'percent'

row += 1
ws[f'A{row}'] = 'Cushion Above 10.5% Buffer (bps)'
ws[f'B{row}'] = '=(C' + str(row-1) + '-10.5)*100'
ws[f'C{row}'] = '=(C' + str(row-1) + '-10.5)*100'
ws[f'D{row}'] = 'bps'

row += 2
ws[f'A{row}'] = 'ASSESSMENT: Base case 5% loss = 40 bps CET1 burn. Cushion remains healthy at ~245 bps.'
ws[f'A{row}'].font = Font(bold=True)
ws[f'A{row}'].fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')

# Bear Case Scenario
row += 3
ws[f'A{row}'] = 'SCENARIO 2: BEAR CASE (15% CUMULATIVE LOSS RATE)'
ws[f'A{row}'].font = Font(bold=True, size=12)
ws[f'A{row}'].fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
row += 1

for col_idx, header in enumerate(headers, start=1):
    cell = ws.cell(row=row, column=col_idx)
    cell.value = header
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color='D4AF37', end_color='D4AF37', fill_type='solid')

row += 1
ws[f'A{row}'] = 'Exposure (M)'
ws[f'B{row}'] = '1,931'
ws[f'C{row}'] = 1931
ws[f'D{row}'] = 'millions'

row += 1
ws[f'A{row}'] = 'Loss Rate (%)'
ws[f'B{row}'] = '15%'
ws[f'C{row}'] = 0.15
ws[f'D{row}'] = 'decimal'

row += 1
ws[f'A{row}'] = 'Expected Loss (M)'
ws[f'B{row}'] = '=C' + str(row-2) + '*C' + str(row-1)
ws[f'C{row}'] = '=C' + str(row-2) + '*C' + str(row-1)
ws[f'D{row}'] = 'millions'

row += 1
ws[f'A{row}'] = 'Tax Rate'
ws[f'B{row}'] = '20%'
ws[f'C{row}'] = 0.20
ws[f'D{row}'] = 'decimal'

row += 1
ws[f'A{row}'] = 'After-Tax Impact (M)'
ws[f'B{row}'] = '=C' + str(row-2) + '*(1-C' + str(row-1) + ')'
ws[f'C{row}'] = '=C' + str(row-2) + '*(1-C' + str(row-1) + ')'
ws[f'D{row}'] = 'millions'

row += 1
ws[f'A{row}'] = 'RWA (M)'
ws[f'B{row}'] = '19,118.5'
ws[f'C{row}'] = 19118.5
ws[f'D{row}'] = 'millions'

row += 1
ws[f'A{row}'] = 'CET1 Burn (bps)'
ws[f'B{row}'] = '=(C' + str(row-2) + '/C' + str(row-1) + ')*10000'
ws[f'C{row}'] = '=(C' + str(row-2) + '/C' + str(row-1) + ')*10000'
ws[f'D{row}'] = 'bps'

row += 1
ws[f'A{row}'] = 'Starting CET1 Ratio (%)'
ws[f'B{row}'] = '13.35'
ws[f'C{row}'] = 13.35
ws[f'D{row}'] = 'percent'

row += 1
ws[f'A{row}'] = 'New CET1 Ratio (%)'
ws[f'B{row}'] = '=C' + str(row-1) + '-(C' + str(row-2) + '/100)'
ws[f'C{row}'] = '=C' + str(row-1) + '-(C' + str(row-2) + '/100)'
ws[f'D{row}'] = 'percent'

row += 1
ws[f'A{row}'] = 'Cushion Above 10.5% Buffer (bps)'
ws[f'B{row}'] = '=(C' + str(row-1) + '-10.5)*100'
ws[f'C{row}'] = '=(C' + str(row-1) + '-10.5)*100'
ws[f'D{row}'] = 'bps'

row += 1
ws[f'A{row}'] = 'Shares Outstanding (M)'
ws[f'B{row}'] = '51.8'
ws[f'C{row}'] = 51.8
ws[f'D{row}'] = 'millions'

row += 1
ws[f'A{row}'] = 'TBVPS Impact ($)'
ws[f'B{row}'] = '=C' + str(row-7) + '/C' + str(row-1)
ws[f'C{row}'] = '=C' + str(row-7) + '/C' + str(row-1)
ws[f'D{row}'] = 'dollars'

row += 1
ws[f'A{row}'] = 'Starting TBVPS ($)'
ws[f'B{row}'] = '36.16'
ws[f'C{row}'] = 36.16
ws[f'D{row}'] = 'dollars'

row += 1
ws[f'A{row}'] = 'New TBVPS ($)'
ws[f'B{row}'] = '=C' + str(row-1) + '-C' + str(row-2)
ws[f'C{row}'] = '=C' + str(row-1) + '-C' + str(row-2)
ws[f'D{row}'] = 'dollars'

row += 1
ws[f'A{row}'] = 'TBVPS Change (%)'
ws[f'B{row}'] = '=(C' + str(row-1) + '/C' + str(row-2) + '-1)*100'
ws[f'C{row}'] = '=(C' + str(row-1) + '/C' + str(row-2) + '-1)*100'
ws[f'D{row}'] = 'percent'

row += 2
ws[f'A{row}'] = 'ASSESSMENT: Bear case 15% loss = 121 bps CET1 burn. Cushion tightens to ~164 bps.'
row += 1
ws[f'A{row}'] = 'Combined with through-cycle NCO normalization (102 bps), total stress = 223 bps burn.'
row += 1
ws[f'A{row}'] = 'This would reduce CET1 to 11.12% (cushion = 62 bps) → MATERIAL CAPITAL CONSTRAINT'
ws[f'A{row}'].font = Font(bold=True)
ws[f'A{row}'].fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')

# Column widths
ws.column_dimensions['A'].width = 35
ws.column_dimensions['B'].width = 25
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 15

# Save workbook
wb.save(workbook_path)

print("✅ Industrial/Warehouse Stress tab added to workbook")
print(f"   File: {workbook_path}")
print(f"   Tab name: Industrial_Warehouse_Stress")
print()
print("Scenarios modeled:")
print("  Base Case (5% loss):  40 bps CET1 burn → Manageable")
print("  Bear Case (15% loss): 121 bps CET1 burn → Material constraint")
print()
print("Combined stress (Bear Industrial + NCO normalization):")
print("  Total burn: 223 bps → CET1 falls to 11.12% (cushion 62 bps)")
